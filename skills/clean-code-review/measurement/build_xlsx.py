#!/usr/bin/env python3
"""Regenerate report.xlsx from the accumulated result TSVs. Idempotent — run after each file.

Raw rows land in data_hits / data_adj / data_llm sheets; every aggregate cell is an Excel
formula (COUNTIFS/SUM/IF/IFERROR) referencing them, so the workbook is auditable and
recomputes if the data sheets are edited.

openpyxl writes formulas without a cached result, which makes the file open blank in any
viewer that does not recalculate (Preview, Quick Look, Numbers). To avoid that, after
saving we inject each formula's computed value as its `<v>` cache — the same thing Excel
does on save. Formulas stay live (fullCalcOnLoad recalculates them in Excel); the cache is
only what non-Excel viewers show.
"""
import csv
import os
import re
import shutil
import zipfile
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
RESULTS = os.path.join(HERE, "results")
MATCH_TOLERANCE_LINES = 3
PROJECTS = ["archon-search", "moonset", "financialwell", "dddd", "udemy-top-down-shooter"]


def read_tsv(name):
    path = os.path.join(RESULTS, name)
    if not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        return [r for r in csv.reader(f, delimiter="\t", quoting=csv.QUOTE_NONE)
                if r and not r[0].startswith("#")]


def inject_cached_values(path, cache, title_to_file):
    """Insert <v>value</v> after <f>…</f> for every cached formula cell, so the file
    displays computed numbers without an Excel recalc. cache: {sheet_title:{ref:value}}."""
    file_cache = {title_to_file[t]: c for t, c in cache.items() if t in title_to_file}
    # openpyxl emits formula cells as <c r=".."><f>…</f><v /></c>; fill that empty <v/>.
    cell_re = re.compile(r'<c r="([A-Z]+\d+)"([^>]*)>(<f[^>]*>.*?</f>)<v\s*/></c>', re.S)
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            name = item.filename.rsplit("/", 1)[-1]
            if item.filename.startswith("xl/worksheets/") and name in file_cache:
                refs = file_cache[name]

                def repl(m, refs=refs):
                    ref, attrs, formula = m.group(1), m.group(2), m.group(3)
                    val = refs.get(ref)
                    if val is None:
                        return m.group(0)
                    return f'<c r="{ref}"{attrs}>{formula}<v>{val}</v></c>'

                data = cell_re.sub(repl, data.decode("utf-8")).encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)


def main():
    catalog = {}  # id -> (severity, title)
    with open(os.path.join(RESULTS, "catalog.txt")) as f:
        for line in f:
            cid, sev, title = [p.strip() for p in line.strip().split("·", 2)]
            catalog[cid] = (sev, title)

    files = []
    with open(os.path.join(HERE, "files.tsv")) as f:
        for r in csv.reader(f, delimiter="\t"):
            if r and not r[0].startswith("#"):
                files.append(r)

    hits = read_tsv("hits.tsv")            # project, file, check, line, excerpt
    adjs = read_tsv("adjudications.tsv")   # project, file, check, line, verdict, reason
    llm = read_tsv("llmonly.tsv")          # project, file, check, line, note

    hit_lines = {}
    file_level_hits = set()  # (check, file) hits that measure the whole file — match by file, not line
    for _p, fp, c, ln, ex, *_ in hits:
        hit_lines.setdefault((c, fp), set()).add(int(ln))
        if ex.startswith("file-level:"):
            file_level_hits.add((c, fp))

    # --- per (check, project) aggregates, used to cache the formula results --
    hits_cp = defaultdict(int)
    for p, _f, c, _ln, _ex in (r[:5] for r in hits):
        hits_cp[(c, p)] += 1
    tp_cp, fp_cp = defaultdict(int), defaultdict(int)
    for p, _f, c, _ln, v, _rs in (r[:6] for r in adjs):
        (tp_cp if v == "TP" else fp_cp)[(c, p)] += 1
    llm_cp, ovl_cp = defaultdict(int), defaultdict(int)
    llm_rows = []
    for p, fpath, c, ln, note in (r[:5] for r in llm):
        matched = ((c, fpath) in file_level_hits or
                   any(abs(int(ln) - h) <= MATCH_TOLERANCE_LINES
                       for h in hit_lines.get((c, fpath), ())))
        llm_cp[(c, p)] += 1
        if matched:
            ovl_cp[(c, p)] += 1
        llm_rows.append([p, fpath, c, int(ln), note, matched])
    hits_by_file = defaultdict(int)
    for _p, fpath, _c, _ln, _ex in (r[:5] for r in hits):
        hits_by_file[fpath] += 1

    wb = Workbook()
    grey = PatternFill("solid", start_color="DDDDDD")
    bold = Font(bold=True)
    cache = {}  # sheet_title -> {cell_ref: computed value} for formula cells

    def data_sheet(name, header, rows):
        ws = wb.create_sheet(name)
        ws.append(header)
        for cell in ws[1]:
            cell.font = bold
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"

    def check_sheet(ws, projects):
        header = (["check", "severity", "title"] + [f"hits:{p}" for p in projects] +
                  ["hits total", "adjudicated", "TP", "FP", "precision %",
                   "LLM-only found", "overlap w/ script", "script misses",
                   "script coverage %", "est. true hits", "quality rate"])
        ws.append(header)
        for cell in ws[1]:
            cell.font = bold
        n = len(projects)
        col = lambda i: chr(ord("A") + i)  # noqa: E731 — 19 cols max, stays in A..Z
        first_hit, total_c = 3, 3 + n
        adj_c, tp_c, fp_c, prec_c = total_c + 1, total_c + 2, total_c + 3, total_c + 4
        llm_c, ovl_c, miss_c = total_c + 5, total_c + 6, total_c + 7
        cov_c, est_c, qual_c = total_c + 8, total_c + 9, total_c + 10
        last_c = qual_c
        refs = cache.setdefault(ws.title, {})
        for i, cid in enumerate(sorted(catalog)):
            row_i = i + 2
            sev, title = catalog[cid]
            row = [cid, sev, title]
            for p in projects:
                row.append(f'=COUNTIFS(data_hits!$A:$A,"{p}",data_hits!$C:$C,$A{row_i})')
            row.append(f"=SUM({col(first_hit)}{row_i}:{col(total_c - 1)}{row_i})")
            row.append(f"={col(tp_c)}{row_i}+{col(fp_c)}{row_i}")
            if n > 1:
                verdict = 'data_adj!$E:$E'
                row.append(f'=COUNTIFS(data_adj!$C:$C,$A{row_i},{verdict},"TP")')
                row.append(f'=COUNTIFS(data_adj!$C:$C,$A{row_i},{verdict},"FP")')
            else:
                p = projects[0]
                row.append(f'=COUNTIFS(data_adj!$A:$A,"{p}",data_adj!$C:$C,$A{row_i},data_adj!$E:$E,"TP")')
                row.append(f'=COUNTIFS(data_adj!$A:$A,"{p}",data_adj!$C:$C,$A{row_i},data_adj!$E:$E,"FP")')
            row.append(f'=IF({col(adj_c)}{row_i}=0,"",ROUND(100*{col(tp_c)}{row_i}/{col(adj_c)}{row_i},1))')
            if n > 1:
                row.append(f'=COUNTIFS(data_llm!$C:$C,$A{row_i})')
                row.append(f'=COUNTIFS(data_llm!$C:$C,$A{row_i},data_llm!$F:$F,TRUE)')
            else:
                p = projects[0]
                row.append(f'=COUNTIFS(data_llm!$A:$A,"{p}",data_llm!$C:$C,$A{row_i})')
                row.append(f'=COUNTIFS(data_llm!$A:$A,"{p}",data_llm!$C:$C,$A{row_i},data_llm!$F:$F,TRUE)')
            row.append(f"={col(llm_c)}{row_i}-{col(ovl_c)}{row_i}")
            row.append(f'=IF({col(llm_c)}{row_i}=0,"",ROUND(100*{col(ovl_c)}{row_i}/{col(llm_c)}{row_i},1))')
            row.append(f'=IF({col(adj_c)}{row_i}=0,"",ROUND({col(total_c)}{row_i}*{col(tp_c)}{row_i}/{col(adj_c)}{row_i},0))')
            # quality rate: TP / LLM-only found — confirmed script haul vs unaided-LLM haul (higher = better)
            row.append(f'=IFERROR({col(tp_c)}{row_i}/{col(llm_c)}{row_i},0)')
            ws.append(row)

            # cache the computed result of every formula cell in this row
            per_project = [hits_cp[(cid, p)] for p in projects]
            total = sum(per_project)
            tp = sum(tp_cp[(cid, p)] for p in projects)
            fp = sum(fp_cp[(cid, p)] for p in projects)
            adj = tp + fp
            found = sum(llm_cp[(cid, p)] for p in projects)
            overlap = sum(ovl_cp[(cid, p)] for p in projects)
            vals = {}
            for k, p in enumerate(projects):
                vals[first_hit + k] = per_project[k]
            vals[total_c] = total
            vals[adj_c] = adj
            vals[tp_c] = tp
            vals[fp_c] = fp
            vals[prec_c] = round(100 * tp / adj, 1) if adj else None
            vals[llm_c] = found
            vals[ovl_c] = overlap
            vals[miss_c] = found - overlap
            vals[cov_c] = round(100 * overlap / found, 1) if found else None
            vals[est_c] = round(total * tp / adj) if adj else None
            vals[qual_c] = tp / found if found else 0
            for cidx, val in vals.items():
                if val is not None:
                    refs[f"{col(cidx)}{row_i}"] = val

        last = ws.max_row
        ws.conditional_formatting.add(
            f"A2:{col(last_c)}{last}",
            FormulaRule(formula=[f"${col(total_c)}2=0"], fill=grey))
        pc = col(prec_c)
        ws.conditional_formatting.add(
            f"{pc}2:{pc}{last}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                           mid_type="num", mid_value=50, mid_color="FFEB84",
                           end_type="num", end_value=100, end_color="63BE7B"))
        ws.freeze_panes = "A2"

    wb.active.title = "Summary"
    check_sheet(wb.active, PROJECTS)
    for p in PROJECTS:
        check_sheet(wb.create_sheet(p[:31]), [p])

    ws = wb.create_sheet("Files")
    ws.append(["project", "path", "lines", "role", "stratum", "script hits", "processed"])
    for cell in ws[1]:
        cell.font = bold
    done = {r[0] for r in read_tsv("progress.tsv")}
    frefs = cache.setdefault("Files", {})
    for i, (proj, path, lines, role, stratum) in enumerate(files):
        row_i = i + 2
        ws.append([proj, path, int(lines), role, stratum,
                   f'=COUNTIFS(data_hits!$B:$B,$B{row_i})',
                   "yes" if path in done else ""])
        frefs[f"F{row_i}"] = hits_by_file[path]
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("Method")
    for row in [
        ["model", "sonnet (adjudication + LLM-only agents)"],
        ["run date", "2026-08-27 (full fresh rerun, live scripts)"],
        ["execution", "collect.sh sweep upfront (all 46 files); then sonnet agents orchestrated by a workflow: adjudicate per packet (<=150 hits, file-boundary split; large single files flushed whole) + LLM-only one agent per file -> assemble TSVs -> rebuild xlsx. Run split across a 3pm session-limit reset (34 agents, then the remaining 27); no data affected."],
        ["adjudication sample", "all hits, every file (no per-check cap)"],
        ["LLM-only granularity", "one agent per file (46 agents), hits withheld — not packet-chunked"],
        ["aggregates", "Excel formulas over data_hits/data_adj/data_llm; results cached at build time so the file displays without a recalc"],
        ["precision", "TP / adjudicated script hits (formula)"],
        ["script coverage %", "overlap / LLM-only found (recall proxy, formula)"],
        ["script misses", "LLM-only found - overlap (formula)"],
        ["est. true hits", "hits total x precision (script haul corrected for FPs, formula)"],
        ["quality rate", "TP / LLM-only found, IFERROR->0 (formula; higher = script confirms more real violations than unaided LLM)"],
        ["LLM-only repeats", "1 (token budget; variance not measured)"],
        ["LLM-only input", "catalog of 60 scriptable checks (id/severity/title), hits withheld"],
        ["match tolerance", f"same check + file, ±{MATCH_TOLERANCE_LINES} lines (data_llm.matched, computed at build time)"],
        ["HIT_CAP", "200 (as shipped, per-file runs, no cap warnings observed unless noted)"],
        ["file selection", "files.tsv (frozen 2026-08-24)"],
    ]:
        ws.append(row)

    data_sheet("data_hits", ["project", "file", "check", "line", "excerpt"],
               [[p, f, c, int(ln), ex] for p, f, c, ln, ex in (r[:5] for r in hits)])
    data_sheet("data_adj", ["project", "file", "check", "line", "verdict", "reason"],
               [[p, f, c, int(ln), v, rs] for p, f, c, ln, v, rs in (r[:6] for r in adjs)])
    data_sheet("data_llm", ["project", "file", "check", "line", "note", "matched"], llm_rows)

    title_to_file = {ws.title: f"sheet{i + 1}.xml" for i, ws in enumerate(wb.worksheets)}
    out = os.path.join(HERE, "report.xlsx")
    wb.save(out)
    inject_cached_values(out, cache, title_to_file)
    print("report.xlsx written:",
          f"{len(hits)} hits, {len(adjs)} adjudications, {len(llm)} llm-only findings")


if __name__ == "__main__":
    main()
